# -*- coding: utf-8 -*-
"""
DB 파일(TSV/CSV/Excel) 파싱 및 FLAG 비트 필터링.
FLAG는 2진 비트 플래그(예: 0x1000 + 0x0100)로 필터링한다.
"""
import csv
import hashlib
import json
import re
import sqlite3
from pathlib import Path
from typing import Optional
from gps_log_generator_app.runtime_paths import get_runtime_root

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

CACHE_SCHEMA_VERSION = "1"


def _project_root() -> Path:
    return get_runtime_root()


def _cache_dir() -> Path:
    d = _project_root() / "data" / "db_cache"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _cache_path_for(source_path: Path) -> Path:
    key = hashlib.sha1(str(source_path.resolve()).encode("utf-8")).hexdigest()
    return _cache_dir() / f"{key}.sqlite3"


def _source_signature(path: Path) -> dict[str, str]:
    st = path.stat()
    return {
        "source_path": str(path.resolve()),
        "source_size": str(st.st_size),
        "source_mtime_ns": str(st.st_mtime_ns),
        "schema_version": CACHE_SCHEMA_VERSION,
    }


def _create_cache_schema(conn: sqlite3.Connection) -> None:
    conn.execute(
        "CREATE TABLE IF NOT EXISTS meta (key TEXT PRIMARY KEY, value TEXT NOT NULL)"
    )
    conn.execute(
        "CREATE TABLE IF NOT EXISTS rows (row_idx INTEGER PRIMARY KEY, row_json TEXT NOT NULL)"
    )


def _read_meta(conn: sqlite3.Connection) -> dict[str, str]:
    cur = conn.execute("SELECT key, value FROM meta")
    return {k: v for k, v in cur.fetchall()}


def _cache_is_valid(conn: sqlite3.Connection, signature: dict[str, str]) -> bool:
    meta = _read_meta(conn)
    return all(meta.get(k) == v for k, v in signature.items())


def _rows_from_source(
    path: Path, delimiter: str, encoding: str, limit: int | None = None, offset: int = 0
) -> list[dict]:
    suf = path.suffix.lower()
    if suf in (".xlsx", ".xls"):
        if suf == ".xls":
            return []
        return _rows_from_excel(path, limit=limit, offset=offset)
    return _rows_from_csv(path, delimiter, encoding, limit=limit, offset=offset)


def _rebuild_cache(
    source_path: Path, cache_path: Path, delimiter: str, encoding: str
) -> bool:
    try:
        rows = _rows_from_source(source_path, delimiter, encoding, limit=None, offset=0)
        signature = _source_signature(source_path)
        tmp_path = cache_path.with_suffix(".tmp")
        if tmp_path.exists():
            tmp_path.unlink()
        conn = sqlite3.connect(str(tmp_path))
        try:
            _create_cache_schema(conn)
            conn.execute("DELETE FROM meta")
            conn.execute("DELETE FROM rows")
            conn.executemany(
                "INSERT INTO meta(key, value) VALUES(?, ?)",
                list(signature.items()),
            )
            conn.executemany(
                "INSERT INTO rows(row_idx, row_json) VALUES(?, ?)",
                [(i, json.dumps(r, ensure_ascii=False)) for i, r in enumerate(rows)],
            )
            conn.commit()
        finally:
            conn.close()
        tmp_path.replace(cache_path)
        return True
    except Exception:
        return False


def _load_rows_from_cache(
    cache_path: Path, limit: int | None = None, offset: int = 0
) -> list[dict]:
    conn = sqlite3.connect(str(cache_path))
    try:
        if limit is None:
            cur = conn.execute(
                "SELECT row_json FROM rows WHERE row_idx >= ? ORDER BY row_idx",
                (offset,),
            )
        else:
            cur = conn.execute(
                "SELECT row_json FROM rows WHERE row_idx >= ? ORDER BY row_idx LIMIT ?",
                (offset, limit),
            )
        return [json.loads(r[0]) for r in cur.fetchall()]
    finally:
        conn.close()


def parse_flag_expression(expr: str) -> Optional[int]:
    """
    '0x1000 + 0x0100' 또는 '0x1000' 형태 문자열을 정수 비트마스크로 변환.
    공백 무시. 빈 문자열이면 None(필터 미사용).
    """
    if not expr or not expr.strip():
        return None
    expr = expr.strip().replace(" ", "")
    if not expr:
        return None
    try:
        total = 0
        for part in re.split(r"\+", expr):
            part = part.strip()
            if not part:
                continue
            if part.startswith("0x") or part.startswith("0X"):
                total |= int(part, 16)
            else:
                total |= int(part, 0)
        return total if total != 0 else None
    except (ValueError, TypeError):
        return None


def _rows_from_csv(
    path: Path, delimiter: str, encoding: str, limit: int | None = None, offset: int = 0
) -> list[dict]:
    """CSV/TSV 파일을 읽어 list[dict] 반환. limit/offset 있으면 해당 구간만."""
    with open(path, "r", encoding=encoding, newline="") as f:
        try:
            sample = f.read(4096)
            f.seek(0)
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample, delimiters="\t,")
        except csv.Error:
            dialect = csv.excel_tab()
        reader = csv.reader(f, dialect)
        raw_rows = list(reader)
    if len(raw_rows) < 2:
        return []
    headers = raw_rows[0]
    seen = set()
    col_indices = [
        (i, h) for i, h in enumerate(headers) if h not in seen and not seen.add(h)
    ]
    data = raw_rows[1:]
    if offset or limit is not None:
        data = data[offset : (offset + limit) if limit is not None else None]
    return [
        {key: (raw[idx] if idx < len(raw) else "") for idx, key in col_indices}
        for raw in data
    ]


def _rows_from_excel(
    path: Path, limit: int | None = None, offset: int = 0
) -> list[dict]:
    """Excel(.xlsx) 첫 시트를 읽어 list[dict] 반환. limit/offset 있으면 해당 구간만(스트리밍)."""
    if not HAS_OPENPYXL:
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    first = next(it, None)
    if first is None:
        wb.close()
        return []
    headers = [str(h).strip() if h is not None else "" for h in first]
    seen = set()
    col_indices = [
        (i, h) for i, h in enumerate(headers) if h not in seen and not seen.add(h)
    ]
    rows = []
    skipped = 0
    for raw in it:
        if skipped < offset:
            skipped += 1
            continue
        row = {}
        for idx, key in col_indices:
            v = raw[idx] if idx < len(raw) else None
            row[key] = "" if v is None else str(v).strip()
        rows.append(row)
        if limit is not None and len(rows) >= limit:
            break
    wb.close()
    return rows


def load_db(
    file_path: str,
    delimiter: str = "\t",
    encoding: str = "utf-8",
    limit: int | None = None,
    offset: int = 0,
) -> list[dict]:
    """
    DB 파일을 읽어 list[dict] 반환.
    limit: 최대 행 수(None이면 전부). offset: 건너뛸 데이터 행 수.
    Excel은 스트리밍으로 offset~offset+limit만 읽음.
    """
    path = Path(file_path)
    if not path.exists():
        return []
    if path.suffix.lower() == ".xls":
        return []

    cache_path = _cache_path_for(path)
    signature = _source_signature(path)
    try:
        if cache_path.exists():
            conn = sqlite3.connect(str(cache_path))
            try:
                _create_cache_schema(conn)
                is_valid = _cache_is_valid(conn, signature)
            finally:
                conn.close()
            if is_valid:
                return _load_rows_from_cache(cache_path, limit=limit, offset=offset)
        if _rebuild_cache(path, cache_path, delimiter, encoding):
            return _load_rows_from_cache(cache_path, limit=limit, offset=offset)
    except Exception:
        pass

    return _rows_from_source(path, delimiter, encoding, limit=limit, offset=offset)


def filter_by_flag(
    rows: list[dict], flag_mask: Optional[int], flag_column: str = "FLAG"
) -> list[dict]:
    """
    FLAG 컬럼 값이 flag_mask와 AND 연산 시 flag_mask와 동일한 행만 반환.
    flag_mask가 None이면 필터 없이 전체 반환.
    """
    if flag_mask is None:
        return rows
    result = []
    for row in rows:
        try:
            val = row.get(flag_column, 0)
            if val == "" or val is None:
                val = 0
            else:
                val = int(val)
            if (val & flag_mask) == flag_mask:
                result.append(row)
        except (ValueError, TypeError):
            continue
    return result


def get_numeric(row: dict, key: str, default: float = 0.0) -> float:
    """행에서 숫자 값 추출."""
    try:
        v = row.get(key, default)
        if v == "" or v is None:
            return default
        return float(v)
    except (ValueError, TypeError):
        return default
